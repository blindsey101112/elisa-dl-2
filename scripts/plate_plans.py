from openpyxl import load_workbook

def get_ods(file):
    wb = load_workbook(file)
    od_ws = wb["Photometric1"]
    ods={}
    for cell in od_ws["H23":"J23"][0]:
        pos[cell.coordinate] = cell.value
    ods["pos"] = pos
    blk = {}
    for cell in od_ws["H24":"M24"][0]:
        blk[cell.coordinate] = cell.value
    ods["blk"] =  blk
    neg = {}
    for cell in od_ws["K23":"M23"][0]:
        neg[cell.coordinate] = cell.value
    ods["neg"] = neg
    sample01_1 = {}
    for cell in od_ws["B17":"G17"][0]:
        sample01_1[cell.coordinate] = cell.value
    ods["sample01_1"] = sample01_1
    sample01_2 = {}
    for cell in od_ws["B18":"G18"][0]:
        sample01_2[cell.coordinate] = cell.value
    ods["sample01_2"] = sample01_2
    sample02_1 = {}
    for cell in od_ws["H17":"M17"][0]:
        sample02_1[cell.coordinate] = cell.value
    ods["sample02_1"] = sample02_1
    sample02_2 = {}
    for cell in od_ws["H18":"M18"][0]:
        sample02_2[cell.coordinate] = cell.value
    ods["sample02_2"] = sample02_2
    sample03_1 = {}
    for cell in od_ws["B19":"G19"][0]:
        sample03_1[cell.coordinate] = cell.value
    ods["sample03_1"] = sample03_1
    sample03_2 = {}
    for cell in od_ws["B20":"G20"][0]:
        sample03_2[cell.coordinate] = cell.value
    ods["sample03_2"] = sample03_2
    sample04_1 = {}
    for cell in od_ws["H19":"M19"][0]:
        sample04_1[cell.coordinate] = cell.value
    ods["sample04_1"] = sample04_1
    sample04_2 = {}
    for cell in od_ws["H20":"M20"][0]:
        sample04_2[cell.coordinate] = cell.value
    ods["sample04_2"] = sample04_2
    sample05_1 = {}
    for cell in od_ws["B21":"G21"][0]:
        sample05_1[cell.coordinate] = cell.value
    ods["sample05_1"] = sample05_1
    sample05_2 = {}
    for cell in od_ws["B22":"G22"][0]:
        sample05_2[cell.coordinate] = cell.value
    ods["sample05_2"] = sample05_2
    sample06_1 = {}
    for cell in od_ws["H21":"M21"][0]:
        sample06_1[cell.coordinate] = cell.value
    ods["sample06_1"] = sample06_1
    sample06_2 = {}
    for cell in od_ws["H22":"M22"][0]:
        sample06_2[cell.coordinate] = cell.value
    ods["sample06_2"] = sample06_2
    sample07_1 = {}
    for cell in od_ws["B23":"G23"][0]:
        sample07_1[cell.coordinate] = cell.value
    ods["sample07_1"] = sample07_1
    sample07_2 = {}
    for cell in od_ws["B24":"G24"][0]:
        sample07_2[cell.coordinate] = cell.value
    ods["sample07_2"] = sample07_2
    return ods


def get_samples(file):
    wb = load_workbook(file)
    sample_ws = wb["PlatePlan"]
    sample_dilution = {}
    sample01 = {}
    for cell in sample_ws["B17":"G17"][0]:
        sample01[cell.coordinate] = cell.value
    sample_dilution["sample01"] = sample01
    sample02 = {}
    for cell in sample_ws["H17":"M17"][0]:
        sample02[cell.coordinate] = cell.value
    sample_dilution["sample02"] = sample02
    sample03 = {}
    for cell in sample_ws["B19":"G19"][0]:
        sample03[cell.coordinate] = cell.value
    sample_dilution["sample03"] = sample03
    sample04 = {}
    for cell in sample_ws["H19":"M19"][0]:
        sample04[cell.coordinate] = cell.value
    sample_dilution["sample04"] = sample04
    sample05 = {}
    for cell in sample_ws["B21":"G21"][0]:
        sample05[cell.coordinate] = cell.value
    sample_dilution["sample05"] = sample05
    sample06 = {}
    for cell in sample_ws["H21":"M21"][0]:
        sample06[cell.coordinate] = cell.value
    sample_dilution["sample06"] = sample06
    sample07 = {}
    for cell in sample_ws["B23":"G23"][0]:
        sample07[cell.coordinate] = cell.value
    sample_dilution["sample07"] = sample07
    return sample_dilution