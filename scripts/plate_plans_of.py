from openpyxl import load_workbook

def get_ods(file):
    wb = load_workbook(file)
    od_ws = wb["Photometric1"]
    ods={}
    std_curve1 = {}
    for cell in od_ws["B23":"M23"][0]:
        std_curve1[cell.coordinate] = cell.value
    ods["std_curve1"] = std_curve1
    std_curve2 = {}
    for cell in od_ws["B24":"M24"][0]:
        std_curve2[cell.coordinate] = cell.value
    ods["std_curve2"] = std_curve2
    pos = {}
    for cell in od_ws["K20":"M20"][0]:
        pos[cell.coordinate] = cell.value
    ods["pos"] = pos
    blk1 = {}
    for cell in od_ws["K22":"M22"][0]:
        blk1[cell.coordinate] = cell.value
    ods["blk1"] =  blk1
    blk2 = {}
    for cell in od_ws["K21":"M21"][0]:
        blk2[cell.coordinate] = cell.value
    ods["blk2"] =  blk2
    ods["neg"] = {"J20": od_ws["J20"].value, "J21": od_ws["J21"].value, "J22": od_ws["J22"].value}
    sample01 = {}
    for cell in od_ws["B17":"C17"][0]:
        sample01[cell.coordinate] = cell.value
    ods["sample01"] = sample01
    sample02 = {}
    for cell in od_ws["B18":"C18"][0]:
        sample02[cell.coordinate] = cell.value
    ods["sample02"] = sample02
    sample03 = {}
    for cell in od_ws["B19":"C19"][0]:
        sample03[cell.coordinate] = cell.value
    ods["sample03"] = sample03
    sample04 = {}
    for cell in od_ws["B20":"C20"][0]:
        sample04[cell.coordinate] = cell.value
    ods["sample04"] = sample04
    sample05 = {}
    for cell in od_ws["B21":"C21"][0]:
        sample05[cell.coordinate] = cell.value
    ods["sample05"] = sample05
    sample06 = {}
    for cell in od_ws["B22":"C22"][0]:
        sample06[cell.coordinate] = cell.value
    ods["sample06"] = sample06
    sample07 = {}
    for cell in od_ws["D17":"E17"][0]:
        sample07[cell.coordinate] = cell.value
    ods["sample07"] = sample07
    sample08 = {}
    for cell in od_ws["D18":"E18"][0]:
        sample08[cell.coordinate] = cell.value
    ods["sample08"] = sample08
    sample09 = {}
    for cell in od_ws["D19":"E19"][0]:
        sample09[cell.coordinate] = cell.value
    ods["sample09"] = sample09
    sample10 = {}
    for cell in od_ws["D20":"E20"][0]:
        sample10[cell.coordinate] = cell.value
    ods["sample10"] = sample10
    sample11 = {}
    for cell in od_ws["D21":"E21"][0]:
        sample11[cell.coordinate] = cell.value
    ods["sample11"] = sample11
    sample12 = {}
    for cell in od_ws["D22":"E22"][0]:
        sample12[cell.coordinate] = cell.value
    ods["sample12"] = sample12
    sample13 = {}
    for cell in od_ws["F17":"G17"][0]:
        sample13[cell.coordinate] = cell.value
    ods["sample13"] = sample13
    sample14 = {}
    for cell in od_ws["F18":"G18"][0]:
        sample14[cell.coordinate] = cell.value
    ods["sample14"] = sample14
    sample15 = {}
    for cell in od_ws["F19":"G19"][0]:
        sample15[cell.coordinate] = cell.value
    ods["sample15"] = sample15
    sample16 = {}
    for cell in od_ws["F20":"G20"][0]:
        sample16[cell.coordinate] = cell.value
    ods["sample16"] = sample16
    sample17 = {}
    for cell in od_ws["F21":"G21"][0]:
        sample17[cell.coordinate] = cell.value
    ods["sample17"] = sample17
    sample18 = {}
    for cell in od_ws["F22":"G22"][0]:
        sample18[cell.coordinate] = cell.value
    ods["sample18"] = sample18
    sample19 = {}
    for cell in od_ws["H17":"I17"][0]:
        sample19[cell.coordinate] = cell.value
    ods["sample19"] = sample19
    sample20 = {}
    for cell in od_ws["H18":"I18"][0]:
        sample20[cell.coordinate] = cell.value
    ods["sample20"] = sample20
    sample21 = {}
    for cell in od_ws["H19":"I19"][0]:
        sample21[cell.coordinate] = cell.value
    ods["sample21"] = sample21
    sample22 = {}
    for cell in od_ws["H20":"I20"][0]:
        sample22[cell.coordinate] = cell.value
    ods["sample22"] = sample22
    sample23 = {}
    for cell in od_ws["H21":"I21"][0]:
        sample23[cell.coordinate] = cell.value
    ods["sample23"] = sample23
    sample24 = {}
    for cell in od_ws["H22":"I22"][0]:
        sample24[cell.coordinate] = cell.value
    ods["sample24"] = sample24
    sample25 = {}
    for cell in od_ws["J17":"K17"][0]:
        sample25[cell.coordinate] = cell.value
    ods["sample25"] = sample25
    sample26 = {}
    for cell in od_ws["J18":"K18"][0]:
        sample26[cell.coordinate] = cell.value
    ods["sample26"] = sample26
    sample27 = {}
    for cell in od_ws["J19":"K19"][0]:
        sample27[cell.coordinate] = cell.value
    ods["sample27"] = sample27
    sample28 = {}
    for cell in od_ws["L17":"M17"][0]:
        sample28[cell.coordinate] = cell.value
    ods["sample28"] = sample28
    sample29 = {}
    for cell in od_ws["L18":"M18"][0]:
        sample29[cell.coordinate] = cell.value
    ods["sample29"] = sample29
    sample30 = {}
    for cell in od_ws["L19":"M19"][0]:
        sample30[cell.coordinate] = cell.value
    ods["sample30"] = sample30

    return ods


def get_samples(file):
    wb = load_workbook(file)
    sample_ws = wb["PlatePlan"]
    sample_dilution = {
               "sample01" : sample_ws["B17"].value,
               "sample02" : sample_ws["B18"].value,
               "sample03" : sample_ws["B19"].value,
               "sample04" : sample_ws["B20"].value,
               "sample05" : sample_ws["B21"].value,
               "sample06": sample_ws["B22"].value,
               "sample07": sample_ws["D17"].value,
               "sample08": sample_ws["D18"].value,
               "sample09": sample_ws["D19"].value,
               "sample10": sample_ws["D20"].value,
               "sample11": sample_ws["D21"].value,
               "sample12": sample_ws["D22"].value,
               "sample13": sample_ws["F17"].value,
               "sample14": sample_ws["F18"].value,
               "sample15": sample_ws["F19"].value,
               "sample16": sample_ws["F20"].value,
               "sample17": sample_ws["F21"].value,
               "sample18": sample_ws["F22"].value,
               "sample19": sample_ws["H17"].value,
               "sample20": sample_ws["H18"].value,
               "sample21": sample_ws["H19"].value,
               "sample22": sample_ws["H20"].value,
               "sample23": sample_ws["H21"].value,
               "sample24": sample_ws["H22"].value,
               "sample25": sample_ws["J17"].value,
               "sample26": sample_ws["J18"].value,
               "sample27": sample_ws["J19"].value,
               "sample28": sample_ws["L17"].value,
               "sample29": sample_ws["L18"].value,
               "sample30": sample_ws["L19"].value
               }
    return sample_dilution